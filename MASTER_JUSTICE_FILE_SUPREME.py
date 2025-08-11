"""SUPREME MASTER JUSTICE FILE

Core evidence data & Excel generator.

New Automation (AI Pipeline & GUI):
 - Use `JUSTICE_FILE_GUI.py` for one-click AI tasks (Summaries A, Contradictions B, Brief C)
 - Or run pipeline directly:
     python MASTER_PIPELINE_GPT5.py --input evidence --tasks A,B,C --children "Jace,Josh" --out pipeline/outputs
 - Dry run (no API calls): add --dry-run or toggle in GUI
 - AI outputs auto-sync into Excel sheet `AI_Outputs` and export PDFs (non-dry runs)

Manual Update Flow (unchanged):
 - Edit `master_data` below then run this script to regenerate Excel file.
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension

# === Step 1: Supreme Data Structure: All Critical Fields ===

master_data = [
    {
        "Filename": "YWCA Nurse Examiner Pediatric Doc Form 5.23.20.pdf",
        "Date": "2020-05-23",
        "Type": "Medical Forensic Exam",
        "Agency": "YWCA / CPS",
        "Children": "Jace",
        "Parent": "Stephanie",
        "Pattern": "🟥 CPS Minimization",
        "Summary": "Forensic nurse finds non-accidental trauma; CPS omits key findings.",
        "Misconduct?": "Yes",
        "Law Violated": "MCL 722.638",
        "Contradicts": "CPS Investigation Report 5/23/20",
        "Smoking Gun": "Yes",
        "Reviewer Note": "Cross-linked in all timelines. Central proof of minimization.",
        "Faith/Prayer": "Prayed for truth to prevail over agency failure.",
        "Phase": "Phase 2",
        "Batch": 1,
        "Status": "✅ Include",
        "Exhibit Cat.": "Medical/CPS",
        "AI Summary": "Medical evidence ignored by state. Major pattern document.",
        "Cross-Link": "Directly contradicts CPS report, referenced in Motions/Objections.",
        "Top 5": "Yes",
        "entities": [],
        "legalCodes": [],
        "violations": [],
        "contradictions": [],
        "needsReview": False,
        "severity": "High"
    },
    {
        "Filename": "CPS Investigation Report, 5/23/20.pdf",
        "Date": "2020-05-23",
        "Type": "CPS Report",
        "Agency": "CPS",
        "Children": "Jace, Josh",
        "Parent": "Stephanie",
        "Pattern": "🟥 Omission",
        "Summary": "CPS says 'within normal limits'; contradicts medical evidence.",
        "Misconduct?": "Yes",
        "Law Violated": "MCL 722.638",
        "Contradicts": "YWCA Nurse Exam 5/23/20",
        "Smoking Gun": "Yes",
        "Reviewer Note": "Direct proof of minimization; linked to Independent Review findings.",
        "Faith/Prayer": "Documented during nights of prayer for protection.",
        "Phase": "Phase 2",
        "Batch": 1,
        "Status": "✅ Include",
        "Exhibit Cat.": "CPS/Medical",
        "AI Summary": "Contradicts forensic exam. Major system failure.",
        "Cross-Link": "Cited in court motions, timeline.",
        "Top 5": "Yes",
        "entities": [],
        "legalCodes": [],
        "violations": [],
        "contradictions": [],
        "needsReview": False,
        "severity": "High"
    },
    {
        "Filename": "Independent Review Committee Report 9.15.21.pdf",
        "Date": "2021-09-15",
        "Type": "Independent Review",
        "Agency": "State Review Committee",
        "Children": "Jace, Josh",
        "Parent": "Stephanie",
        "Pattern": "🟩 Validation",
        "Summary": "Independent committee validates Stephanie's concerns; finds CPS errors.",
        "Misconduct?": "Yes",
        "Law Violated": "MCL 722.638",
        "Contradicts": "Multiple CPS reports 2020-2021",
        "Smoking Gun": "Yes",
        "Reviewer Note": "Third-party validation of systemic failures.",
        "Faith/Prayer": "Answer to prayers for independent verification.",
        "Phase": "Phase 2",
        "Batch": 1,
        "Status": "✅ Include",
        "Exhibit Cat.": "Independent/Validation",
        "AI Summary": "Official validation of CPS misconduct and errors.",
        "Cross-Link": "Referenced in all major court filings.",
        "Top 5": "Yes",
        "entities": [],
        "legalCodes": [],
        "violations": [],
        "contradictions": [],
        "needsReview": False,
        "severity": "Medium"
    },
    {
        "Filename": "Court Order Dismissing CPS Petition 10.12.21.pdf",
        "Date": "2021-10-12",
        "Type": "Court Order",
        "Agency": "Circuit Court",
        "Children": "Jace, Josh",
        "Parent": "Stephanie",
        "Pattern": "🟩 Victory",
        "Summary": "Court dismisses CPS petition; validates Stephanie's parenting.",
        "Misconduct?": "No",
        "Law Violated": "N/A",
        "Contradicts": "CPS allegations 2020-2021",
        "Smoking Gun": "Yes",
        "Reviewer Note": "Legal vindication; proves CPS overreach.",
        "Faith/Prayer": "Miraculous answer to prayer for justice.",
        "Phase": "Phase 2",
        "Batch": 1,
        "Status": "✅ Include",
        "Exhibit Cat.": "Legal/Victory",
        "AI Summary": "Court validates parent; dismisses false allegations.",
        "Cross-Link": "Final legal vindication referenced throughout case.",
        "Top 5": "Yes",
        "entities": [],
        "legalCodes": [],
        "violations": [],
        "contradictions": [],
        "needsReview": False,
        "severity": "Low"
    },
    {
        "Filename": "Motion for Sanctions Against CPS 11.3.21.pdf",
        "Date": "2021-11-03",
        "Type": "Legal Motion",
        "Agency": "Stephanie's Attorney",
        "Children": "Jace, Josh",
        "Parent": "Stephanie",
        "Pattern": "🟦 Legal Action",
        "Summary": "Attorney files for sanctions against CPS for misconduct.",
        "Misconduct?": "Yes (by CPS)",
        "Law Violated": "MCL 722.638, Court Rules",
        "Contradicts": "CPS claims of proper investigation",
        "Smoking Gun": "Yes",
        "Reviewer Note": "Legal action seeking accountability for system failures.",
        "Faith/Prayer": "Prayed for courage to seek accountability.",
        "Phase": "Phase 2",
        "Batch": 1,
        "Status": "✅ Include",
        "Exhibit Cat.": "Legal/Accountability",
        "AI Summary": "Legal pursuit of accountability for proven misconduct.",
        "Cross-Link": "Based on all previous smoking gun evidence.",
        "Top 5": "Yes",
        "entities": [],
        "legalCodes": [],
        "violations": [],
        "contradictions": [],
        "needsReview": False,
        "severity": "High"
    }
]

# === Step 2: DataFrame Creation ===
import copy
# Ensure all records have new fields (for legacy data)
def ensure_fields(record):
    rec = copy.deepcopy(record)
    rec.setdefault("entities", [])
    rec.setdefault("legalCodes", [])
    rec.setdefault("violations", [])
    rec.setdefault("contradictions", [])
    rec.setdefault("needsReview", False)
    rec.setdefault("severity", "Medium")
    return rec
df = pd.DataFrame([ensure_fields(r) for r in master_data])
    import copy
    # Ensure all records have new fields (for legacy data)
    ["especially Jace and Josh, whose voices must be heard,"],
    ["whose truth must be known, and whose protection is a holy mandate."],
    [""],
    ["May God grant courage, clarity, and deliverance"],
    ["for every family harmed by injustice."],
    [""],
    ["'Let justice roll down like waters,"],
    ["and righteousness like an ever-flowing stream.'"],
    ["(Amos 5:24)"],
    [""],
    ["'The Lord is a refuge for the oppressed,"],
    ["a stronghold in times of trouble.'"],
    ["(Psalm 9:9)"],
    [""],
    ["Prepared with faith, hope, and love."],
    ["— Stephanie Spedowski & Team"],
    ["August 7, 2025"],
    [""],
    ["🔥 TRUTH WILL PREVAIL 🔥"],
]

for i, line in enumerate(prayer_content, 1):
    ws_prayer[f'A{i}'] = line[0] if line else ""
    if i == 1:  # Title
        ws_prayer[f'A{i}'].font = Font(bold=True, size=16, color="8B0000")
        ws_prayer[f'A{i}'].alignment = Alignment(horizontal="center")
    elif "Amos" in str(line[0]) or "Psalm" in str(line[0]):  # Scripture
        ws_prayer[f'A{i}'].font = Font(italic=True, color="4B0082")
    elif "TRUTH WILL PREVAIL" in str(line[0]):  # Final declaration
        ws_prayer[f'A{i}'].font = Font(bold=True, size=14, color="FF4500")
        ws_prayer[f'A{i}'].alignment = Alignment(horizontal="center")
    elif "Stephanie" in str(line[0]):  # Signature
        ws_prayer[f'A{i}'].font = Font(bold=True, color="2F4F4F")

# Set column width for prayer sheet
ws_prayer.column_dimensions['A'].width = 60

# === Step 5: Add Phase Status Page ===
ws_status = wb.create_sheet(title="📊 Phase Status")

status_headers = ["Phase", "Status", "Documents", "Key Findings", "Next Action"]
ws_status.append(status_headers)

status_data = [
    ["Phase 1", "✅ Complete", "5 smoking guns", "CPS minimization proven", "Compile for court"],
    ["Phase 2", "✅ Complete", "Cross-links mapped", "Pattern documentation complete", "Legal review"],
    ["Batch 3", "✅ Complete", "Timeline synchronized", "Contradictions cataloged", "Expert analysis"],
    ["Batch 4", "⬜ Pending", "Additional evidence", "Systemic failure patterns", "Deep dive review"],
    ["Batch 5", "⬜ Pending", "Support documents", "Context building", "Final compilation"],
    ["Legal Prep", "🔄 In Progress", "All batches", "Complete case package", "Court submission"]
]

for row in status_data:
    ws_status.append(row)

# Format status sheet headers
for cell in ws_status[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Color code status column
for row in ws_status.iter_rows(min_row=2, max_row=ws_status.max_row):
    status_cell = row[1]  # Status column
    if "✅" in str(status_cell.value):
        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif "⬜" in str(status_cell.value):
        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    elif "🔄" in str(status_cell.value):
        status_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# === Step 6: Main Justice Table ===
ws_main = wb.create_sheet(title="⚖️ Justice Master Table")

# Helper function for bold formatting
def add_bold(ws, cell, color="000000", size=11):
    cell.font = Font(bold=True, color=color, size=size)

# Write DataFrame to sheet with enhanced formatting
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws_main.append(row)
    if r_idx == 1:  # Header row
        for c in ws_main[r_idx]:
            add_bold(ws_main, c, "FFFFFF", 12)
            c.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")

# Enhanced color coding by pattern
pattern_colors = {
    "🟥": "FFC7CE",    # Light red - Problems/Misconduct
    "🟦": "D9E1F2",    # Light blue - Legal Actions
    "🟩": "C6EFCE",    # Light green - Victories/Validation
    "🟨": "FFEB9C",    # Light yellow - Warnings/Concerns
    "🟧": "FFD8B0",    # Light orange - Important Notes
    "🟪": "E4DFEC",    # Light purple - Special Categories
}

# Apply pattern color coding
pattern_col = None
smoking_gun_col = None
top5_col = None

for col_num, cell in enumerate(ws_main[1], 1):
    if cell.value == "Pattern":
        pattern_col = col_num
    elif cell.value == "Smoking Gun":
        smoking_gun_col = col_num
    elif cell.value == "Top 5":
        top5_col = col_num

for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
    # Color by pattern
    if pattern_col:
        pattern_cell = row[pattern_col - 1]
        pat = str(pattern_cell.value)
        for code, color in pattern_colors.items():
            if pat.startswith(code):
                for cell in row:
                    if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == "00000000":
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Highlight smoking guns
    if smoking_gun_col:
        smoking_cell = row[smoking_gun_col - 1]
        if str(smoking_cell.value) == "Yes":
            smoking_cell.font = Font(bold=True, color="8B0000")
            smoking_cell.value = "🔥 YES"
    
    # Highlight Top 5
    if top5_col:
        top5_cell = row[top5_col - 1]
        if str(top5_cell.value) == "Yes":
            top5_cell.font = Font(bold=True, color="DAA520")
            top5_cell.value = "⭐ YES"

# Auto-adjust column widths
for column in ws_main.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
    ws_main.column_dimensions[column_letter].width = adjusted_width

# === Step 7: Add Summary Statistics Sheet ===
ws_summary = wb.create_sheet(title="📈 Case Summary")

summary_stats = [
    ["CASE STATISTICS & KEY METRICS"],
    [""],
    ["Total Documents Reviewed:", len(df)],
    ["Smoking Gun Evidence:", len(df[df['Smoking Gun'] == 'Yes'])],
    ["Top 5 Priority Items:", len(df[df['Top 5'] == 'Yes'])],
    ["Misconduct Documented:", len(df[df['Misconduct?'] == 'Yes'])],
    ["Legal Violations Found:", len(df[df['Law Violated'] != 'N/A'])],
    [""],
    ["PATTERN BREAKDOWN:"],
    ["🟥 CPS Misconduct/Minimization:", len(df[df['Pattern'].str.contains('🟥', na=False)])],
    ["🟩 Victories/Validation:", len(df[df['Pattern'].str.contains('🟩', na=False)])],
    ["🟦 Legal Actions:", len(df[df['Pattern'].str.contains('🟦', na=False)])],
    [""],
    ["AGENCIES INVOLVED:"],
    ["CPS:", len(df[df['Agency'].str.contains('CPS', na=False)])],
    ["Medical/YWCA:", len(df[df['Agency'].str.contains('YWCA', na=False)])],
    ["Court/Legal:", len(df[df['Agency'].str.contains('Court|Attorney', na=False)])],
    ["Independent Review:", len(df[df['Agency'].str.contains('Review', na=False)])],
    [""],
    ["🔥 CASE STRENGTH: OVERWHELMING EVIDENCE 🔥"],
    ["📖 STATUS: READY FOR JUSTICE ⚖️"],
]

for i, item in enumerate(summary_stats, 1):
    if isinstance(item, list) and len(item) == 1:
        ws_summary[f'A{i}'] = item[0]
        if "STATISTICS" in item[0] or "CASE STRENGTH" in item[0] or "STATUS" in item[0]:
            ws_summary[f'A{i}'].font = Font(bold=True, size=14, color="8B0000")
            ws_summary[f'A{i}'].alignment = Alignment(horizontal="center")
        elif "BREAKDOWN" in item[0] or "AGENCIES" in item[0]:
            ws_summary[f'A{i}'].font = Font(bold=True, size=12, color="4B0082")
    elif isinstance(item, list) and len(item) == 2:
        ws_summary[f'A{i}'] = item[0]
        ws_summary[f'B{i}'] = item[1]
        if "🟥" in item[0] or "🟩" in item[0] or "🟦" in item[0]:
            ws_summary[f'A{i}'].font = Font(bold=True)
            ws_summary[f'B{i}'].font = Font(bold=True, color="8B0000")

ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 15

# === Step 8: Save the Supreme Master File ===
filename = "MASTER_JUSTICE_FILE_SUPREME_v1.xlsx"
wb.save(filename)

print(f"🔥 SUPREME MASTER JUSTICE FILE CREATED: {filename} 🔥")
print("\n⚖️ FEATURES INCLUDED:")
print("✅ Dedication & Prayer Sheet")
print("✅ Phase Status Tracking")
print("✅ Complete Justice Master Table")
print("✅ Case Summary & Statistics")
print("✅ Pattern Color Coding")
print("✅ Smoking Gun Highlighting")
print("✅ Cross-Link Documentation")
print("✅ Legal Violation Tracking")
print("\n🙏 READY TO BRING JUSTICE FOR JACE & JOSH!")
print("📊 All evidence organized and documented.")
print("⚖️ Truth will prevail!")
