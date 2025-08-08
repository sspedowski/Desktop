"""Excel synchronization utilities.

Appends parsed AI output metadata into MASTER_JUSTICE_FILE_SUPREME_v1.xlsx
Sheet: AI_Outputs (created if missing)

Captured fields per markdown summary file (Task A focus):
- Timestamp
- Base Filename
- Tasks Run
- Children
- Include (Yes/No/UNKNOWN)
- Smoking Gun (Yes/No/UNKNOWN)
- Top 5 (Yes/No/UNKNOWN)
- Primary Pattern
- Tags (raw list)
- Markdown Path
- PDF Path (if exists)
"""
from __future__ import annotations
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional
from openpyxl import load_workbook, Workbook

SUMMARY_PATTERNS = [
    ('include', [
        re.compile(r'^Include\??.*?:\s*(.*)$', re.IGNORECASE),
        re.compile(r'^Include Decision\s*:?\s*(.*)$', re.IGNORECASE)
    ]),
    ('smoking', [
        re.compile(r'^Smoking Gun\??.*?:\s*(.*)$', re.IGNORECASE),
        re.compile(r'^Is Smoking Gun\s*:?\s*(.*)$', re.IGNORECASE)
    ]),
    ('top5', [
        re.compile(r'^Top\s*5\??.*?:\s*(.*)$', re.IGNORECASE),
        re.compile(r'^Top Five\s*:?\s*(.*)$', re.IGNORECASE)
    ]),
    ('primary_pattern', [
        re.compile(r'^Primary Pattern\s*:?\s*(.*)$', re.IGNORECASE),
        re.compile(r'^Key Pattern\s*:?\s*(.*)$', re.IGNORECASE)
    ]),
    ('tags', [
        re.compile(r'^Tags?\s*:?\s*\[(.*?)]\s*$', re.IGNORECASE),
        re.compile(r'^Tags?\s*:?\s*(.+)$', re.IGNORECASE)
    ]),
]

HEADERS = [
    'Timestamp','Filename','Tasks','Children','Include','Smoking Gun','Top 5',
    'Primary Pattern','Tags','Markdown Path','PDF Path'
]

def parse_summary(md_text: str) -> Dict[str,str]:
    data: Dict[str,str] = {}
    for raw_line in md_text.splitlines():
        line = raw_line.strip()
        if not line or len(line) > 500:
            continue
        for key, patterns in SUMMARY_PATTERNS:
            if key in data:
                continue
            for rx in patterns:
                m = rx.match(line)
                if m:
                    captured = m.group(1).strip()
                    # Trim trailing markdown artifacts
                    captured = captured.rstrip('#*').strip()
                    data[key] = captured or 'UNKNOWN'
                    break
    # Normalize values
    def norm_yes_no(v: Optional[str]) -> str:
        if not v:
            return 'UNKNOWN'
        v2 = v.lower()
        if any(tok in v2 for tok in ['yes','y']):
            return 'Yes'
        if any(tok in v2 for tok in ['no','n']):
            return 'No'
        return 'UNKNOWN'
    data['include'] = norm_yes_no(data.get('include'))
    data['smoking'] = norm_yes_no(data.get('smoking'))
    data['top5'] = norm_yes_no(data.get('top5'))
    # Normalize tags: split by comma if not already bracket list format
    if 'tags' in data and data['tags'] and '[' not in data['tags']:
        parts = [p.strip() for p in re.split(r'[;,]', data['tags']) if p.strip()]
        if parts:
            data['tags'] = ', '.join(parts)
    return data

def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        try:
            return load_workbook(path)
        except Exception:
            pass
    wb = Workbook()
    wb.active.title = 'AI_Outputs'
    ws = wb['AI_Outputs']
    ws.append(HEADERS)
    return wb

def sync_markdowns(md_dir: Path, excel_path: Path, tasks: str, children: str, pdf_root: Path | None) -> int:
    md_files = list(md_dir.glob('*.md'))
    if not md_files:
        return 0
    wb = ensure_workbook(excel_path)
    if 'AI_Outputs' not in wb.sheetnames:
        ws = wb.create_sheet('AI_Outputs')
        ws.append(HEADERS)
    ws = wb['AI_Outputs']
    existing = {row[1].value for row in ws.iter_rows(min_row=2, values_only=False)}  # existing filenames
    added = 0
    for md in md_files:
        base_name = md.stem
        display_name = base_name.replace('SUMMARY_','') + md.suffix
        if display_name in existing:
            # skip duplicates of same summary
            continue
        content = md.read_text(encoding='utf-8', errors='ignore')
        parsed = parse_summary(content)
    pdf_path = (pdf_root / (md.stem + '.pdf')) if pdf_root else None
    row = [
            datetime.utcnow().isoformat(timespec='seconds')+'Z',
            display_name,
            tasks,
            children,
            parsed.get('include','UNKNOWN'),
            parsed.get('smoking','UNKNOWN'),
            parsed.get('top5','UNKNOWN'),
            parsed.get('primary_pattern','UNKNOWN'),
            parsed.get('tags',''),
            str(md),
            str(pdf_path) if (pdf_path and pdf_path.exists()) else ''
        ]
        ws.append(row)
        added += 1
    wb.save(excel_path)
    return added
